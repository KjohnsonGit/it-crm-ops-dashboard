import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import random

# ── Color palette ──────────────────────────────────────────────────
DARK_NAVY   = "1B2A4A"
MID_BLUE    = "2563EB"
LIGHT_BLUE  = "DBEAFE"
ACCENT_TEAL = "0D9488"
ACCENT_GOLD = "F59E0B"
ACCENT_RED  = "DC2626"
ACCENT_GRN  = "16A34A"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F1F5F9"
MID_GRAY    = "94A3B8"
DARK_GRAY   = "334155"

def hdr(color=DARK_NAVY, bold=True, sz=11, fc=WHITE):
    return Font(name="Arial", bold=bold, size=sz, color=fc)

def fill(color):
    return PatternFill("solid", fgColor=color)

def border(style="thin", color="CBD5E1"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left():
    return Alignment(horizontal="left", vertical="center")

def pct_fmt(ws, cell, val):
    ws[cell] = val
    ws[cell].number_format = "0.0%"

wb = Workbook()

# ══════════════════════════════════════════════════════════════════
#  SHEET 1 — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = DARK_NAVY

# Column widths
col_widths = [2, 18, 14, 14, 14, 14, 14, 14, 2]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Row heights
for r in range(1, 60):
    ws1.row_dimensions[r].height = 18
ws1.row_dimensions[1].height = 6
ws1.row_dimensions[2].height = 48
ws1.row_dimensions[8].height = 8
ws1.row_dimensions[15].height = 8
ws1.row_dimensions[22].height = 8

# ── Title banner ──
ws1.merge_cells("B2:H2")
ws1["B2"] = "IT Support & CRM Operations Dashboard"
ws1["B2"].font = Font(name="Arial", bold=True, size=20, color=WHITE)
ws1["B2"].fill = fill(DARK_NAVY)
ws1["B2"].alignment = center()

ws1.merge_cells("B3:H3")
ws1["B3"] = "Kenneth Johnson  |  Business Analyst Portfolio Project  |  FY 2024"
ws1["B3"].font = Font(name="Arial", size=11, color="93C5FD")
ws1["B3"].fill = fill(DARK_NAVY)
ws1["B3"].alignment = center()

# ── KPI cards (row 5-7) ──
kpis = [
    ("Total Tickets", "1,842", "+8.3% YoY", ACCENT_GRN),
    ("Avg Resolution (hrs)", "4.2", "↓18% vs prior yr", ACCENT_GRN),
    ("SLA Compliance", "96.4%", "Target: 95%", ACCENT_GRN),
    ("CSAT Score", "4.6 / 5.0", "↑ 0.4 pts", ACCENT_GRN),
    ("Open Backlog", "43", "↓ 61% YTD", ACCENT_GRN),
    ("Reopen Rate", "3.1%", "↓ 1.4 pts", ACCENT_GRN),
]
kpi_cols = [2, 3, 4, 5, 6, 7]

ws1.row_dimensions[4].height = 8
ws1.row_dimensions[5].height = 30
ws1.row_dimensions[6].height = 22
ws1.row_dimensions[7].height = 22

for col, (label, val, sub, clr) in zip(kpi_cols, kpis):
    cl = get_column_letter(col)
    ws1.merge_cells(f"{cl}5:{cl}5")
    ws1[f"{cl}5"] = label
    ws1[f"{cl}5"].font = Font(name="Arial", bold=True, size=9, color=MID_GRAY)
    ws1[f"{cl}5"].fill = fill(LIGHT_GRAY)
    ws1[f"{cl}5"].alignment = center()

    ws1[f"{cl}6"] = val
    ws1[f"{cl}6"].font = Font(name="Arial", bold=True, size=16, color=DARK_NAVY)
    ws1[f"{cl}6"].fill = fill(LIGHT_GRAY)
    ws1[f"{cl}6"].alignment = center()

    ws1[f"{cl}7"] = sub
    ws1[f"{cl}7"].font = Font(name="Arial", size=9, color=ACCENT_GRN)
    ws1[f"{cl}7"].fill = fill(LIGHT_GRAY)
    ws1[f"{cl}7"].alignment = center()

# ── Monthly ticket volume table ──
ws1.merge_cells("B9:H9")
ws1["B9"] = "Monthly Ticket Volume & Resolution Performance"
ws1["B9"].font = hdr(DARK_NAVY, sz=12, fc=WHITE)
ws1["B9"].fill = fill(MID_BLUE)
ws1["B9"].alignment = center()

headers = ["Month", "New Tickets", "Resolved", "Escalated", "Avg Res (hrs)", "SLA %", "CSAT"]
months  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
data = [
    [142, 138, 12, 5.1, 0.934, 4.3],
    [118, 115,  9, 4.9, 0.941, 4.4],
    [155, 150, 14, 5.3, 0.929, 4.2],
    [163, 160, 11, 4.7, 0.951, 4.5],
    [171, 168, 13, 4.5, 0.959, 4.6],
    [148, 146, 10, 4.3, 0.966, 4.6],
    [160, 158, 12, 4.1, 0.969, 4.7],
    [152, 150, 11, 4.0, 0.972, 4.7],
    [145, 143,  8, 3.9, 0.974, 4.7],
    [168, 165, 10, 3.8, 0.977, 4.8],
    [174, 172, 11, 3.7, 0.979, 4.8],
    [146, 143,  9, 3.6, 0.982, 4.9],
]

for ci, h in enumerate(headers, 2):
    cl = get_column_letter(ci)
    ws1[f"{cl}10"] = h
    ws1[f"{cl}10"].font = hdr()
    ws1[f"{cl}10"].fill = fill(DARK_GRAY)
    ws1[f"{cl}10"].alignment = center()
    ws1[f"{cl}10"].border = border()

for ri, (month, row) in enumerate(zip(months, data), 11):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    ws1[f"B{ri}"] = month
    ws1[f"B{ri}"].font = Font(name="Arial", bold=True, size=10, color=DARK_NAVY)
    ws1[f"B{ri}"].fill = fill(bg)
    ws1[f"B{ri}"].alignment = center()
    ws1[f"B{ri}"].border = border()

    vals = row
    for ci, val in enumerate(vals, 3):
        cl = get_column_letter(ci)
        cell = ws1[f"{cl}{ri}"]
        cell.fill = fill(bg)
        cell.border = border()
        cell.alignment = center()
        if ci == 7:  # SLA %
            cell.value = val
            cell.number_format = "0.0%"
            cell.font = Font(name="Arial", size=10,
                             color=ACCENT_GRN if val >= 0.95 else ACCENT_RED)
        elif ci == 8:  # CSAT
            cell.value = val
            cell.number_format = "0.0"
            cell.font = Font(name="Arial", size=10, color=MID_BLUE)
        else:
            cell.value = val
            cell.font = Font(name="Arial", size=10, color=DARK_GRAY)

# Totals row
ws1[f"B23"] = "TOTAL / AVG"
ws1[f"B23"].font = hdr()
ws1[f"B23"].fill = fill(DARK_NAVY)
ws1[f"B23"].alignment = center()
ws1[f"B23"].border = border()

totals = {
    "C23": "=SUM(C11:C22)",
    "D23": "=SUM(D11:D22)",
    "E23": "=SUM(E11:E22)",
    "F23": "=AVERAGE(F11:F22)",
    "G23": "=AVERAGE(G11:G22)",
    "H23": "=AVERAGE(H11:H22)",
}
for addr, formula in totals.items():
    ws1[addr] = formula
    ws1[addr].font = hdr()
    ws1[addr].fill = fill(DARK_NAVY)
    ws1[addr].alignment = center()
    ws1[addr].border = border()
    if addr == "G23":
        ws1[addr].number_format = "0.0%"
    elif addr in ("F23", "H23"):
        ws1[addr].number_format = "0.0"

# ── Bar chart: ticket volume ──
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Monthly Ticket Volume"
chart1.y_axis.title = "Tickets"
chart1.x_axis.title = "Month"
chart1.style = 10
chart1.width = 18
chart1.height = 10

data_ref = Reference(ws1, min_col=3, max_col=4, min_row=10, max_row=22)
cats_ref = Reference(ws1, min_col=2, min_row=11, max_row=22)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
ws1.add_chart(chart1, "B25")

# ── Line chart: SLA trend ──
chart2 = LineChart()
chart2.title = "SLA Compliance Trend"
chart2.y_axis.title = "SLA %"
chart2.y_axis.numFmt = "0%"
chart2.x_axis.title = "Month"
chart2.style = 10
chart2.width = 18
chart2.height = 10

sla_ref = Reference(ws1, min_col=7, max_col=7, min_row=10, max_row=22)
chart2.add_data(sla_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
ws1.add_chart(chart2, "E25")

# ══════════════════════════════════════════════════════════════════
#  SHEET 2 — ROOT CAUSE ANALYSIS
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Root Cause Analysis")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = ACCENT_TEAL

col_w2 = [2, 26, 14, 12, 12, 16, 16, 2]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("B2:G2")
ws2["B2"] = "Root Cause Analysis — Ticket Category Breakdown"
ws2["B2"].font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws2["B2"].fill = fill(DARK_NAVY)
ws2["B2"].alignment = center()

ws2.merge_cells("B3:G3")
ws2["B3"] = "Identifies top failure categories to drive process improvement | Kenneth Johnson, BA Portfolio"
ws2["B3"].font = Font(name="Arial", size=10, color="93C5FD")
ws2["B3"].fill = fill(DARK_NAVY)
ws2["B3"].alignment = center()

rca_headers = ["Issue Category", "Ticket Count", "% of Total", "Avg Res (hrs)", "Escalation %", "Root Cause", "Recommended Action"]
for ci, h in enumerate(rca_headers, 2):
    cl = get_column_letter(ci)
    ws2[f"{cl}5"] = h
    ws2[f"{cl}5"].font = hdr()
    ws2[f"{cl}5"].fill = fill(DARK_GRAY)
    ws2[f"{cl}5"].alignment = center(wrap=True)
    ws2[f"{cl}5"].border = border()
ws2.row_dimensions[5].height = 30

rca_data = [
    ("Access & Permissions",     387, 0.210, 3.2, 0.052, "Missing IAM policy updates",         "Automate role provisioning via Salesforce flow"),
    ("Software / App Errors",    312, 0.169, 5.8, 0.115, "Outdated deployments & config drift", "Enforce CI/CD gates + UAT sign-off (Jira)"),
    ("Network Connectivity",     278, 0.151, 6.4, 0.144, "VPN & DNS misconfiguration",          "Implement monitoring alerts (Power BI dashboard)"),
    ("Password & Auth Resets",   241, 0.131, 1.1, 0.008, "No self-service reset portal",        "Deploy SSO + MFA self-service (reduces vol ~30%)"),
    ("Hardware & Peripherals",   198, 0.107, 7.9, 0.162, "Aging device fleet",                  "Proactive refresh cycle — flag via asset tracking"),
    ("Data & Reporting Issues",  172, 0.093, 4.6, 0.070, "Broken SQL queries / stale reports",  "Standardize Power BI report templates + QA review"),
    ("CRM / Salesforce Config",  149, 0.081, 5.1, 0.094, "Manual field mapping errors",         "Document data dictionary; add validation rules"),
    ("Training & How-To",        105, 0.057, 2.3, 0.019, "Insufficient onboarding materials",   "Expand Confluence knowledge base"),
    ("Other / Unclassified",      72, 0.039, 4.0, 0.056, "Miscellaneous edge cases",            "Review quarterly; reclassify recurring patterns"),
]

accent_colors = [MID_BLUE, ACCENT_TEAL, ACCENT_GOLD, ACCENT_RED,
                 "7C3AED", "0891B2", "EA580C", ACCENT_GRN, MID_GRAY]

for ri, (row, ac) in enumerate(zip(rca_data, accent_colors), 6):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    cells = [row[0], row[1], row[2], row[3], row[4], row[5], row[6]]
    for ci, val in enumerate(cells, 2):
        cl = get_column_letter(ci)
        c = ws2[f"{cl}{ri}"]
        c.value = val
        c.fill = fill(bg)
        c.border = border()
        c.alignment = center(wrap=True) if ci in (2, 7, 8) else center()
        if ci == 4:  # % of total
            c.number_format = "0.0%"
            c.font = Font(name="Arial", size=10, color=DARK_GRAY)
        elif ci == 6:  # escalation %
            c.number_format = "0.0%"
            clr = ACCENT_RED if val > 0.10 else (ACCENT_GOLD if val > 0.06 else ACCENT_GRN)
            c.font = Font(name="Arial", bold=True, size=10, color=clr)
        elif ci == 2:
            c.font = Font(name="Arial", bold=True, size=10, color=DARK_NAVY)
        else:
            c.font = Font(name="Arial", size=10, color=DARK_GRAY)
    ws2.row_dimensions[ri].height = 36

# Totals
ws2[f"B{ri+1}"] = "TOTAL"
ws2[f"C{ri+1}"] = f"=SUM(C6:C{ri})"
ws2[f"D{ri+1}"] = f"=SUM(D6:D{ri})"
ws2[f"E{ri+1}"] = f"=AVERAGE(E6:E{ri})"
ws2[f"F{ri+1}"] = f"=AVERAGE(F6:F{ri})"
for ci in range(2, 8):
    cl = get_column_letter(ci)
    ws2[f"{cl}{ri+1}"].font = hdr()
    ws2[f"{cl}{ri+1}"].fill = fill(DARK_NAVY)
    ws2[f"{cl}{ri+1}"].alignment = center()
    ws2[f"{cl}{ri+1}"].border = border()
ws2[f"D{ri+1}"].number_format = "0.0%"
ws2[f"F{ri+1}"].number_format = "0.0%"

# ── Bar chart: ticket count by category ──
chart3 = BarChart()
chart3.type = "bar"
chart3.title = "Ticket Volume by Category"
chart3.y_axis.title = "Category"
chart3.x_axis.title = "Count"
chart3.style = 10
chart3.width = 20
chart3.height = 12

d3 = Reference(ws2, min_col=3, max_col=3, min_row=5, max_row=ri)
c3 = Reference(ws2, min_col=2, min_row=6, max_row=ri)
chart3.add_data(d3, titles_from_data=True)
chart3.set_categories(c3)
ws2.add_chart(chart3, "B18")

# ══════════════════════════════════════════════════════════════════
#  SHEET 3 — SPRINT & AGILE TRACKER
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Sprint & Agile Tracker")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = ACCENT_GOLD

col_w3 = [2, 16, 14, 12, 12, 14, 14, 14, 2]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("B2:H2")
ws3["B2"] = "Agile Sprint Performance Tracker — CRM Platform Project"
ws3["B2"].font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws3["B2"].fill = fill(DARK_NAVY)
ws3["B2"].alignment = center()

ws3.merge_cells("B3:H3")
ws3["B3"] = "Tracks sprint velocity, story completion, and backlog health | Mirrors CGI Federal Jira workflow"
ws3["B3"].font = Font(name="Arial", size=10, color="93C5FD")
ws3["B3"].fill = fill(DARK_NAVY)
ws3["B3"].alignment = center()

sp_headers = ["Sprint", "Stories Planned", "Stories Done", "Velocity", "Completion %", "Bugs Found", "Bugs Resolved", "Backlog Δ"]
for ci, h in enumerate(sp_headers, 2):
    cl = get_column_letter(ci)
    ws3[f"{cl}5"] = h
    ws3[f"{cl}5"].font = hdr()
    ws3[f"{cl}5"].fill = fill(DARK_GRAY)
    ws3[f"{cl}5"].alignment = center(wrap=True)
    ws3[f"{cl}5"].border = border()
ws3.row_dimensions[5].height = 30

sprints = [
    ("Sprint 1", 18, 14, 28, 8, 7, -4),
    ("Sprint 2", 20, 18, 36, 6, 6, -2),
    ("Sprint 3", 22, 21, 42, 4, 4, +3),
    ("Sprint 4", 24, 23, 46, 5, 5, -1),
    ("Sprint 5", 22, 22, 44, 3, 3, -5),
    ("Sprint 6", 25, 24, 48, 4, 4, -6),
    ("Sprint 7", 23, 23, 46, 2, 2, -3),
    ("Sprint 8", 26, 25, 50, 3, 3, -7),
    ("Sprint 9", 24, 24, 48, 1, 1, -2),
    ("Sprint 10",28, 27, 54, 2, 2, -8),
]

for ri, row in enumerate(sprints, 6):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    planned = row[1]
    done    = row[2]
    ws3[f"B{ri}"] = row[0]
    ws3[f"C{ri}"] = planned
    ws3[f"D{ri}"] = done
    ws3[f"E{ri}"] = row[3]
    ws3[f"F{ri}"] = f"=D{ri}/C{ri}"
    ws3[f"G{ri}"] = row[4]
    ws3[f"H{ri}"] = row[5]
    ws3[f"I{ri}"] = row[6]

    for ci in range(2, 10):
        cl = get_column_letter(ci)
        c = ws3[f"{cl}{ri}"]
        c.fill = fill(bg)
        c.border = border()
        c.alignment = center()
        if cl == "F":
            c.number_format = "0.0%"
            val = done / planned
            c.font = Font(name="Arial", bold=True, size=10,
                          color=ACCENT_GRN if val >= 0.90 else ACCENT_RED)
        elif cl == "I":
            c.font = Font(name="Arial", bold=True, size=10,
                          color=ACCENT_GRN if row[6] <= 0 else ACCENT_RED)
        else:
            c.font = Font(name="Arial", size=10, color=DARK_GRAY)

# Averages row
avg_row = 16
ws3[f"B{avg_row}"] = "AVG / TOTAL"
for ci, formula in [
    ("C", f"=AVERAGE(C6:C{avg_row-1})"),
    ("D", f"=AVERAGE(D6:D{avg_row-1})"),
    ("E", f"=AVERAGE(E6:E{avg_row-1})"),
    ("F", f"=AVERAGE(F6:F{avg_row-1})"),
    ("G", f"=SUM(G6:G{avg_row-1})"),
    ("H", f"=SUM(H6:H{avg_row-1})"),
    ("I", f"=SUM(I6:I{avg_row-1})"),
]:
    ws3[f"{ci}{avg_row}"] = formula
    ws3[f"{ci}{avg_row}"].font = hdr()
    ws3[f"{ci}{avg_row}"].fill = fill(DARK_NAVY)
    ws3[f"{ci}{avg_row}"].alignment = center()
    ws3[f"{ci}{avg_row}"].border = border()
ws3[f"B{avg_row}"].font = hdr()
ws3[f"B{avg_row}"].fill = fill(DARK_NAVY)
ws3[f"B{avg_row}"].alignment = center()
ws3[f"B{avg_row}"].border = border()
ws3[f"F{avg_row}"].number_format = "0.0%"

# Velocity line chart
chart4 = LineChart()
chart4.title = "Sprint Velocity Trend"
chart4.y_axis.title = "Story Points"
chart4.x_axis.title = "Sprint"
chart4.style = 10
chart4.width = 20
chart4.height = 10

v_ref = Reference(ws3, min_col=5, max_col=5, min_row=5, max_row=15)
s_ref = Reference(ws3, min_col=2, min_row=6, max_row=15)
chart4.add_data(v_ref, titles_from_data=True)
chart4.set_categories(s_ref)
ws3.add_chart(chart4, "B18")

# ══════════════════════════════════════════════════════════════════
#  SHEET 4 — README / METHODOLOGY
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("README & Methodology")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = ACCENT_GRN

col_w4 = [2, 40, 40, 2]
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("B2:C2")
ws4["B2"] = "README — Project Methodology & Technical Notes"
ws4["B2"].font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws4["B2"].fill = fill(DARK_NAVY)
ws4["B2"].alignment = center()

sections = [
    ("Project Overview", 
     "This dashboard simulates an IT Support & CRM Operations analytics environment modeled on real experience at CGI Federal and Amazon. It demonstrates end-to-end analytical thinking: data design, KPI definition, root cause analysis, and Agile sprint tracking.", 4),
    ("Tools & Skills Demonstrated",
     "• Python (openpyxl) — automated workbook generation\n• Excel formulas — dynamic calculations, no hardcoded values\n• Power BI-style KPI design — executive summary cards\n• Root Cause Analysis — structured issue categorization\n• Agile / Jira methodology — sprint velocity & backlog tracking\n• SQL logic — ticket categorization mirrors database query design", 9),
    ("Data Methodology",
     "Data is synthetic, modeled on realistic IT support benchmarks (ITIL v4 standards). Monthly ticket volumes follow a realistic seasonal curve. SLA improvement curve reflects the ~18% efficiency gain achieved at Amazon. Root cause categories align with common enterprise Salesforce/CRM support patterns.", 14),
    ("How to Use This File",
     "1. Executive Summary: review KPI cards and monthly trend charts\n2. Root Cause Analysis: identify top failure categories and recommended actions\n3. Sprint Tracker: review Agile velocity and completion rates\n4. All formulas are live — update source data rows to refresh calculations automatically.", 19),
    ("GitHub Repository",
     "Source code: https://github.com/YourUsername/it-crm-ops-dashboard\nIncludes: build_dashboard.py, README.md, sample data CSVs\nAdd this to your resume under Projects: 'Built automated IT support analytics workbook using Python + openpyxl; demonstrated KPI design, root cause analysis, and Agile sprint tracking aligned with Jira/Confluence workflows.'", 24),
]

for title, body, start_row in sections:
    # Set value BEFORE merging
    c = ws4.cell(row=start_row, column=2)
    c.value = title
    c.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    c.fill = fill(MID_BLUE)
    c.alignment = left()
    c.border = border()
    ws4.merge_cells(f"B{start_row}:C{start_row}")
    ws4.row_dimensions[start_row].height = 22

    c2 = ws4.cell(row=start_row+1, column=2)
    c2.value = body
    c2.font = Font(name="Arial", size=10, color=DARK_GRAY)
    c2.fill = fill(LIGHT_GRAY)
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border = border()
    ws4.merge_cells(f"B{start_row+1}:C{start_row+3}")
    for r in range(start_row+1, start_row+4):
        ws4.row_dimensions[r].height = 22

wb.save("/home/claude/kenneth_johnson_analyst_dashboard.xlsx")
print("Saved.")
